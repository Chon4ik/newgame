from random import *
from pygame import *
import sys
from openpyxl import load_workbook
class GameSprite(sprite.Sprite):
    def __init__(self,color,x,y,w,h,spd):
        super().__init__()
        self.image = Surface((w,h))
        self.image.fill(color)
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y
        self.speed = spd
    def reset(self):
        window.blit(self.image,(self.rect.x,self.rect.y))
class Player(GameSprite):
    def __init__(self,color,x,y,w,h,spd):
        super().__init__(color,x,y,w,h,spd)
        self.vel_y = 0
        self.on_ground = True
    def update(self):
        keys = key.get_pressed() 
        if keys[K_a] and self.rect.x > 5:
            self.rect.x -= self.speed 
        if keys[K_d] and self.rect.x  <win_width-50:
            self.rect.x += self.speed
        if keys[K_SPACE] and self.on_ground:
            self.vel_y = -13
        self.on_ground = False
        self.vel_y += 0.6
        self.rect.y += int(self.vel_y)
        for p in platforms:
            if self.rect.colliderect(p.rect):
                if self.vel_y > 0:
                    self.rect.bottom = p.rect.top
                    self.vel_y = 0
                    self.on_ground = True
                elif self.vel_y < 0:
                    self.rect.top = p.rect.bottom
                    self.vel_y = 0
class Coin(sprite.Sprite):
    def __init__(self,x,y):
        super().__init__()
        self.image = Surface((20,20),SRCALPHA)
        draw.circle(self.image, (255,215,0), (10,10),10)
        self.rect = self.image.get_rect()
        self.rect.center = (x,y)
class Enemy(GameSprite):
    def __init__(self,color,x,y,w,h,spd,left,right):
        super().__init__(color,x,y,w,h,spd)
        self.left_bound = left
        self.right_bound = right
    def update(self):
        self.rect.x += self.speed
        if self.rect.right >= self.right_bound or self.rect.left <= self.left_bound:
            self.speed *= -1
class Flag(GameSprite):
    def __init__(self,x,y):
        super().__init__((255,140,0),x,y,40,80,0)
def load_level(index):
    global platforms,enemys,flag
    data = level_data[index]
    platforms = sprite.Group()
    for x,y,w in data['platforms']:
        platforms.add(GameSprite((60,179,60),x,y,w,20,0))
    coins = sprite.Group()
    for x,y in data['coins']:
        coins.add(Coin(x,y))
    enemys = sprite.Group()
    for x,y,l,r in data['enemys']:
        enemys.add(Enemy((110,43,83),x,y,45,45,3,l,r))
    flag = sprite.GroupSingle(Flag(*data['flag']))
    player1.rect.x = data['start'][0]
    player1.rect.y = data['start'][1]
    player1.vel_y = 0
level_data = [
    {
        'bg': (135, 206, 235),
        'start': (60, 400),
        'platforms': [(0, 460, 800), (150, 360, 150), (350, 280, 150), (550, 200, 150), (100, 180, 120)],
        'coins':     [(230, 330), (425, 250), (625, 170), (160, 150), (400, 430)],
        'enemys':   [(150, 324, 150, 300), (360, 244, 350, 500)],
        'flag':      (730, 400),
    },
    {
        'bg': (135, 206, 235),
        'start': (30, 400),
        'platforms': [(0, 460, 800), (50, 380, 100), (200, 300, 100), (350, 220, 100), (500, 300, 100), (650, 380, 150)],
        'coins':     [(100, 355), (250, 275), (400, 195), (550, 275), (700, 355)],
        'enemys':   [(55, 344, 50, 150), (355, 184, 350, 450), (505, 264, 500, 600)],
        'flag':      (750, 400),
    },
]

CELL = 40
COLOR_MAP = {
    "60B33C": "platform",
    "FFD700": "coin",
    "FF0000": "enemy",
    "FFC000": "flag",
}
def load_level_from_xlsx(path):
    global level_data, current_level
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    platforms = [] 
    coins = []
    enemys = []
    flag_pos = (0, 0)

    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if not (fill and fill.fill_type == "solid"):
                continue
            color = fill.fgColor.rgb[-6:].upper()
            kind = COLOR_MAP.get(color)

            x = (cell.column - 1) * CELL
            y = (cell.row - 1) * CELL

            if kind == "platform":
                platforms.append((x, y, CELL))  
            elif kind == "coin":
                coins.append((x, y))
            elif kind == "enemy":
                enemys.append((x, y, max(0, x - 3*CELL), x + 3*CELL))
            elif kind == "flag":
                flag_pos = (x, y)

    # Добавляем новый уровень в список
    level_data.append({
        'bg': (135, 206, 235),
        'start': (CELL, win_height - 3*CELL),
        'platforms': platforms,
        'coins': coins,
        'enemys': enemys,
        'flag': flag_pos
    })
    current_level = len(level_data) - 1
    load_level(current_level)
font.init()
font1 = font.Font(None,36)
mixer.init()
window = display.set_mode((700,500))
win_width = 700
win_height = 500
speed1 = 5
x1 = 50
y1= 405
speed_y = 6
speed2 = 6
display.set_caption('Платформер')
flag = Flag(650,370)
player1=Player((50,50,255),x1,y1,45,45,speed1)
enemy1 = Enemy((110,43,83),340,305,45,45,3,250,450)
enemy2 = Enemy((110,43,83),460,205,45,45,3,450,650)
enemy3 = Enemy((110,43,83),200,155,45,45,3,100,300)
enemys = sprite.Group()
enemy1.add(enemys)
enemy2.add(enemys)
enemy3.add(enemys)
coin1 = Coin(350,335)
coin2 = Coin(550,235)
coin3 = Coin(200,185)
coins = sprite.Group()
coin1.add(coins)
coin2.add(coins)
coin3.add(coins)
platform = GameSprite((60,179,60),0,450,700,20,0)
platform1 = GameSprite((60,179,60),250,350,200,20,0)
platform2 = GameSprite((60,179,60),450,250,200,20,0)
platform3 = GameSprite((60,179,60),100,200,200,20,0)
platforms = sprite.Group()
platforms.add(platform)
platforms.add(platform1)
platforms.add(platform2)
platforms.add(platform3)
background = transform.scale(image.load('galaxy.jpg'),(700,500))
clock = time.Clock()
lose = font1.render('you lose',True,(255,0,0))
text_win = font1.render('you win',True,(255,255,55))
total = 0
FPS = 60
total = 0
life = 3
finish = False
game = True
flagg=sprite.GroupSingle()
flagg.add(flag)
if len(sys.argv) > 1:
    load_level_from_xlsx(sys.argv[1])   
else:
    current_level = 0
    load_level(current_level)
clock = time.Clock()
while game:
    for e in event.get():
            if e.type == QUIT:
                game = False
    if finish!=True:
        sprites_list1 = sprite.spritecollide(player1,enemys,False)
        sprites_list2 = sprite.spritecollide(player1,coins,True)
        window.blit(background,(0,0))
        platforms.draw(window)
        player1.reset()
        player1.update()
        coins.draw(window)
        enemys.draw(window)
        enemys.update()
        flag.draw(window)
        if sprites_list2:
            total +=1
        if total==3 and sprite.spritecollide(player1,flagg,False):
            window.blit(text_win,(300,250))
            finish = True
        if sprites_list1:
            window.blit(lose,(300,250))
            finish = True
    display.update()
    clock.tick(FPS)